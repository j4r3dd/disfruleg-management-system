# -*- coding: utf-8 -*-
"""
Excel Export Service - VERSIÓN OPTIMIZADA PARA 30 PRODUCTOS
Formato: DISFRULEG (basado en Numbers)
"""

import os
from datetime import datetime
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.modules.receipts.services.base_service import BaseService
from src.modules.receipts.components.user_preferences import get_preferences
from src.modules.receipts.logging_config import get_logger, PerformanceLogger, LogContext
from src.modules.receipts.exceptions import (
    ExcelGenerationError,
    FileWriteError,
    DirectoryNotFoundError,
)
import logging


logger = get_logger(__name__)


BUSINESS_CONFIG = {
    "name": "DISFRULEG",
    "subtitle": "Comercializadora Castruita",
    "address": "Felipe Páramo #160, Veinte de Noviembre C.P. 58219, Morelia, Michoacán",
    "phone": "Cel. (443) 504 9098",
    "rfc": "CACJ850827UF3",
    "regimen_fiscal": "Régimen Simplificado de Confianza",
}

COLOR_HEADER_DARK = 'FF8B1A1A'
COLOR_SECTION_HEADER = 'FFC62828'
COLOR_TEXT_DARK = 'FF000000'


class ExcelExportService(BaseService):
    """Service for Excel generation with enhanced customization"""

    def __init__(self):
        super().__init__()
        self.preferences = get_preferences()
        self.company_name = BUSINESS_CONFIG["name"]
        self.logger = logger
        self.perf = PerformanceLogger(logger)

    def _get_export_directory(self) -> str:
        """Get Excel export directory from user preferences"""
        excel_dir = self.preferences.get_or_prompt_export_path('excel')

        if not excel_dir:
            self._log_warning("Usuario canceló selección, usando directorio predeterminado")
            project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
            excel_dir = os.path.join(project_root, "output", "excel")

        os.makedirs(excel_dir, exist_ok=True)
        return excel_dir

    def _generate_filename(
        self,
        client_name: str,
        folio: Optional[int] = None,
        with_sections: bool = False
    ) -> str:
        """Generate Excel filename: DDMMYY_NOMBRECLIENTE.xlsx"""
        fecha = datetime.now().strftime("%d%m%y")
        safe_name = client_name.replace(' ', '_')
        return f"{fecha}_{safe_name}.xlsx"

    def _get_border(self) -> Border:
        """Get standard border"""
        thin = Side(style='thin', color='000000')
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _add_company_header(self, ws, current_row: int) -> int:
        """Add company header"""
        border = self._get_border()
        
        # Company name
        ws.merge_cells(f'A{current_row}:E{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = BUSINESS_CONFIG["name"]
        cell.font = Font(name='Helvetica', size=14, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLOR_HEADER_DARK, end_color=COLOR_HEADER_DARK, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        # Subtitle
        ws.merge_cells(f'A{current_row}:E{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = BUSINESS_CONFIG["subtitle"]
        cell.font = Font(name='Helvetica', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.row_dimensions[current_row].height = 18
        current_row += 1
        
        # Address
        ws.merge_cells(f'A{current_row}:E{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = BUSINESS_CONFIG["address"]
        cell.font = Font(name='Helvetica', size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.row_dimensions[current_row].height = 16
        current_row += 1
        
        # Phone and RFC
        ws.merge_cells(f'A{current_row}:E{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"{BUSINESS_CONFIG['phone']}"
        cell.font = Font(name='Helvetica', size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.row_dimensions[current_row].height = 16
        current_row += 2
        
        return current_row

    def _add_section_header(self, ws, current_row: int, section_name: str) -> int:
        """Add section header with name"""
        border = self._get_border()
        
        ws.merge_cells(f'A{current_row}:E{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"✦ {section_name.upper()} ✦"
        cell.font = Font(name='Helvetica', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLOR_SECTION_HEADER, end_color=COLOR_SECTION_HEADER, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
        return current_row

    def _add_table_header(self, ws, current_row: int) -> int:
        """Add table header: PRODUCTO, UNIDAD, CANTIDAD, P.UNITARIO, TOTAL"""
        border = self._get_border()
        headers = ['PRODUCTO', 'UNIDAD', 'CANTIDAD', 'P.UNITARIO', 'TOTAL']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = Font(name='Helvetica', size=9, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=COLOR_HEADER_DARK, end_color=COLOR_HEADER_DARK, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        ws.row_dimensions[current_row].height = 18
        return current_row + 1

    def _add_items(self, ws, current_row: int, items: List[List[str]]) -> tuple[int, float]:
        """Add items to table - Orden: PRODUCTO, UNIDAD, CANTIDAD, P.UNITARIO, TOTAL"""
        border = self._get_border()
        total = 0.0
        
        for item in items:
            producto, cantidad_str, unidad, precio_str, subtotal_str = item
            
            # Column 1: PRODUCTO
            ws.cell(row=current_row, column=1).value = producto
            ws.cell(row=current_row, column=1).font = Font(name='Helvetica', size=8)
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=current_row, column=1).border = border
            
            # Column 2: UNIDAD
            ws.cell(row=current_row, column=2).value = unidad
            ws.cell(row=current_row, column=2).font = Font(name='Helvetica', size=8)
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row, column=2).border = border
            
            # Column 3: CANTIDAD
            ws.cell(row=current_row, column=3).value = float(cantidad_str)
            ws.cell(row=current_row, column=3).font = Font(name='Helvetica', size=8)
            ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row, column=3).border = border
            ws.cell(row=current_row, column=3).number_format = '0.00'
            
            # Column 4: P.UNITARIO
            precio = float(precio_str.replace('$', '').replace(',', ''))
            ws.cell(row=current_row, column=4).value = precio
            ws.cell(row=current_row, column=4).font = Font(name='Helvetica', size=8)
            ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='right', vertical='center')
            ws.cell(row=current_row, column=4).border = border
            ws.cell(row=current_row, column=4).number_format = '$#,##0.00'
            
            # Column 5: TOTAL
            subtotal = float(subtotal_str.replace('$', '').replace(',', ''))
            ws.cell(row=current_row, column=5).value = subtotal
            ws.cell(row=current_row, column=5).font = Font(name='Helvetica', size=8)
            ws.cell(row=current_row, column=5).alignment = Alignment(horizontal='right', vertical='center')
            ws.cell(row=current_row, column=5).border = border
            ws.cell(row=current_row, column=5).number_format = '$#,##0.00'
            
            ws.row_dimensions[current_row].height = 16
            total += subtotal
            current_row += 1
        
        return current_row, total

    def _add_subtotal(self, ws, current_row: int, label: str, subtotal: float) -> int:
        """Add subtotal row"""
        border = self._get_border()
        
        # Merge and add label
        ws.merge_cells(f'A{current_row}:D{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = label
        cell.font = Font(name='Helvetica', size=8, bold=True)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        
        # Add value
        cell = ws.cell(row=current_row, column=5)
        cell.value = subtotal
        cell.font = Font(name='Helvetica', size=8, bold=True)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        cell.number_format = '$#,##0.00'
        
        ws.row_dimensions[current_row].height = 16
        return current_row + 1

    def generate_simple_excel(
        self,
        client_name: str,
        items: List[List[str]],
        total: float,
        folio: Optional[int] = None
    ) -> Optional[str]:
        """Generate simple Excel"""
        operation_name = f"Generate simple Excel | {client_name}"

        with LogContext(self.logger, operation_name, logging.INFO):
            try:
                num_items = len(items)
                self.logger.info(
                    f"Excel request | {client_name} | "
                    f"Items: {num_items} | Total: ${total:.2f}"
                )

                export_dir = self._get_export_directory()
                filename = self._generate_filename(client_name, folio)
                file_path = os.path.join(export_dir, filename)

                self.perf.start_timer("excel_creation")
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Pedido"

                # Set column widths
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 10
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 12

                current_row = 1
                current_row = self._add_company_header(ws, current_row)
                current_row = self._add_section_header(ws, current_row, "GENERAL")
                current_row = self._add_table_header(ws, current_row)
                current_row, _ = self._add_items(ws, current_row, items)
                current_row = self._add_subtotal(ws, current_row, "Subtotal GENERAL:", total)

                wb.save(file_path)
                wb.close()
                excel_time = self.perf.end_timer("excel_creation", logging.INFO)

                file_size = os.path.getsize(file_path)
                self.logger.info(f"✅ Excel generated | {file_path} | {file_size} bytes | Config: {num_items} items")
                return file_path

            except Exception as e:
                self.logger.error(f"Excel generation failed | {client_name}", exc_info=True)
                raise ExcelGenerationError(client_name=client_name, reason=str(e))

    def generate_sectioned_excel(
        self,
        client_name: str,
        sections: Dict[str, Dict[str, Any]],
        total: float,
        folio: Optional[int] = None
    ) -> Optional[str]:
        """Generate Excel with sections"""
        operation_name = f"Generate sectioned Excel | {client_name}"

        with LogContext(self.logger, operation_name, logging.INFO):
            try:
                total_items = sum(len(s['items']) for s in sections.values())
                self.logger.info(
                    f"Sectioned Excel | {client_name} | "
                    f"Sections: {len(sections)} | Items: {total_items}"
                )

                export_dir = self._get_export_directory()
                filename = self._generate_filename(client_name, folio, with_sections=True)
                file_path = os.path.join(export_dir, filename)

                self.perf.start_timer("excel_creation")
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Pedido"

                # Set column widths
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 10
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 12

                current_row = 1
                current_row = self._add_company_header(ws, current_row)

                # Process each section
                for idx, (section_name, section_data) in enumerate(sections.items(), 1):
                    self.logger.debug(
                        f"Processing section {idx}/{len(sections)}: {section_name} "
                        f"({len(section_data['items'])} items)"
                    )

                    current_row = self._add_section_header(ws, current_row, section_name)
                    current_row = self._add_table_header(ws, current_row)
                    
                    items_section = section_data['items']
                    current_row, _ = self._add_items(ws, current_row, items_section)
                    
                    subtotal_section = section_data['subtotal']
                    current_row = self._add_subtotal(ws, current_row, f"Subtotal {section_name}:", subtotal_section)
                    current_row += 1

                wb.save(file_path)
                wb.close()
                excel_time = self.perf.end_timer("excel_creation", logging.INFO)

                file_size = os.path.getsize(file_path)
                self.logger.info(f"✅ Sectioned Excel | {file_path} | {file_size} bytes | Config: {total_items} items")
                return file_path

            except Exception as e:
                self.logger.error(f"Sectioned Excel failed | {client_name}", exc_info=True)
                raise ExcelGenerationError(
                    client_name=client_name,
                    reason=f"Sectioned Excel error: {str(e)}"
                )

    def generate_automatic_excel(
        self,
        client_name: str,
        items_carrito: Optional[List[List[str]]] = None,
        items_por_seccion: Optional[Dict[str, Dict[str, Any]]] = None,
        total: Optional[float] = None,
        folio: Optional[int] = None
    ) -> Optional[str]:
        """Generate Excel automatically"""
        try:
            if items_por_seccion and len(items_por_seccion) > 1:
                if isinstance(total, str):
                    total_float = float(total.replace('$', '').replace(',', ''))
                else:
                    total_float = float(total)

                return self.generate_sectioned_excel(client_name, items_por_seccion, total_float, folio)
            else:
                if items_carrito is None and items_por_seccion:
                    items_carrito = []
                    for datos_seccion in items_por_seccion.values():
                        items_carrito.extend(datos_seccion['items'])

                if isinstance(total, str):
                    total_float = float(total.replace('$', '').replace(',', ''))
                else:
                    total_float = float(total)

                return self.generate_simple_excel(client_name, items_carrito, total_float, folio)

        except Exception as e:
            self._log_error(f"Error al generar Excel automático: {e}", e)
            return None

    @staticmethod
    def convert_sections_to_simple(items_por_seccion: Dict[str, Dict[str, Any]]) -> List[List[str]]:
        """Convert sectioned data to simple format"""
        items_simple = []
        for datos_seccion in items_por_seccion.values():
            items_simple.extend(datos_seccion['items'])
        return items_simple