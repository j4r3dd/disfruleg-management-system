# -*- coding: utf-8 -*-
"""
Export Service - Analytics Module
Business logic for exporting analytics data to various formats
Sigue Clean Architecture con PDFs profesionales y consistentes
"""
from typing import Dict, List, Optional, Tuple
from datetime import datetime
import logging
import os

logger = logging.getLogger(__name__)

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False


class ExportService:
    """
    Servicio para exportar datos de analytics a Excel y PDF
    Todos los PDFs usan el mismo estilo profesional y corporativo
    """

    # Colores corporativos DISFRULEG
    COLOR_PRIMARY = colors.HexColor("#2D5A27")      # Verde bosque
    COLOR_TEXT = colors.HexColor("#333333")         # Gris oscuro
    COLOR_BG = colors.HexColor("#F4F4F4")          # Gris claro
    COLOR_ACCENT = colors.HexColor("#E8F0E5")      # Verde muy claro

    def __init__(self, assets_path: str = None):
        """Inicializar servicio de exportación"""
        self.assets_path = assets_path or os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 
            "assets"
        )

    def _get_logo_path(self) -> Optional[str]:
        """Obtener ruta del logo si existe"""
        logo_path = os.path.join(self.assets_path, "logo_disfruleg.jpg")
        return logo_path if os.path.exists(logo_path) else None

    def export_product_detail_to_excel(
        self,
        product_data: Dict,
        detail_data: Dict,
        filepath: str
    ) -> Tuple[bool, str]:
        """
        Exportar detalle de producto a Excel
        """
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            logger.info(f"Exporting product detail to Excel: {filepath}")

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Hoja 1: Información del producto
                info_data = {
                    'Campo': [
                        'ID Producto',
                        'Nombre',
                        'Unidad',
                        'Cantidad Vendida',
                        'Ingresos Totales',
                        'Ganancia Total',
                        'Margen %',
                        'Stock Actual',
                        'Costos Totales',
                        'Fecha Exportación'
                    ],
                    'Valor': [
                        product_data.get('id_producto', 'N/A'),
                        product_data.get('nombre_producto', 'N/A'),
                        product_data.get('unidad_producto', 'N/A'),
                        f"{float(product_data.get('cantidad_vendida', 0)):,.2f}",
                        f"${float(product_data.get('ingresos_totales', 0)):,.2f}",
                        f"${float(product_data.get('ganancia_total', 0)):,.2f}",
                        f"{float(product_data.get('margen_ganancia_porcentaje', 0)):.2f}%",
                        f"{float(product_data.get('stock', 0)):,.2f}",
                        f"${float(product_data.get('costos_totales', 0)):,.2f}",
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    ]
                }
                df_info = pd.DataFrame(info_data)
                df_info.to_excel(writer, sheet_name='Información', index=False)

                # Hoja 2: Historial de ventas
                sales_data = detail_data.get('historial_ventas', [])
                if sales_data:
                    df_sales = pd.DataFrame(sales_data)
                    df_sales.to_excel(writer, sheet_name='Historial Ventas', index=False)

                # Hoja 3: Top clientes
                clients_data = detail_data.get('top_clientes', [])
                if clients_data:
                    df_clients = pd.DataFrame(clients_data)
                    df_clients.to_excel(writer, sheet_name='Top Clientes', index=False)

            # Aplicar formato
            self._format_excel_workbook(filepath)

            logger.info(f"✅ Product detail exported successfully to {filepath}")
            return True, f"Archivo exportado exitosamente a:\n{filepath}"

        except ImportError:
            error_msg = "Se requiere pandas y openpyxl para exportar a Excel.\n\nInstala con: pip install pandas openpyxl"
            logger.error(error_msg)
            return False, error_msg
        except Exception as e:
            error_msg = f"Error al exportar a Excel: {str(e)}"
            logger.error(f"❌ {error_msg}", exc_info=True)
            return False, error_msg

    def export_client_detail_to_excel(
        self,
        client_data: Dict,
        detail_data: Dict,
        filepath: str,
        period: str = "30D"
    ) -> Tuple[bool, str]:
        """
        Exportar detalle de cliente a Excel
        """
        try:
            import pandas as pd

            logger.info(f"Exporting client detail to Excel: {filepath}")

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Hoja 1: Información del cliente
                info_data = {
                    'Campo': [
                        'ID Cliente',
                        'Nombre',
                        'Grupo',
                        'Tipo',
                        'Periodo',
                        'Total Compras',
                        'Notas Generadas',
                        'Saldo Pendiente',
                        'Descuento',
                        'Última Compra',
                        'Fecha Exportación'
                    ],
                    'Valor': [
                        client_data.get('id_cliente', 'N/A'),
                        client_data.get('nombre_cliente', 'N/A'),
                        client_data.get('clave_grupo', 'N/A'),
                        client_data.get('tipo_cliente', 'N/A'),
                        period,
                        f"${float(client_data.get('total_ventas', 0)):,.2f}",
                        f"{int(client_data.get('cantidad_facturas', 0)):,}",
                        f"${float(client_data.get('saldo_pendiente', 0)):,.2f}",
                        f"{float(client_data.get('porcentaje_descuento', 0)):.1f}%",
                        self._format_date(client_data.get('ultima_compra')),
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    ]
                }
                df_info = pd.DataFrame(info_data)
                df_info.to_excel(writer, sheet_name='Información', index=False)

                # Hoja 2: Historial de compras
                purchases_data = detail_data.get('historial_compras', [])
                if purchases_data:
                    df_purchases = pd.DataFrame(purchases_data)
                    df_purchases.to_excel(writer, sheet_name='Historial Compras', index=False)

                # Hoja 3: Top productos
                products_data = detail_data.get('top_productos', [])
                if products_data:
                    df_products = pd.DataFrame(products_data)
                    df_products.to_excel(writer, sheet_name='Top Productos', index=False)

                # Hoja 4: Deudas
                debts_data = detail_data.get('deudas', [])
                if debts_data:
                    df_debts = pd.DataFrame(debts_data)
                    df_debts.to_excel(writer, sheet_name='Deudas', index=False)

            # Aplicar formato
            self._format_excel_workbook(filepath)

            logger.info(f"✅ Client detail exported successfully to {filepath}")
            return True, f"Archivo exportado exitosamente a:\n{filepath}"

        except ImportError:
            error_msg = "Se requiere pandas y openpyxl para exportar a Excel.\n\nInstala con: pip install pandas openpyxl"
            logger.error(error_msg)
            return False, error_msg
        except Exception as e:
            error_msg = f"Error al exportar a Excel: {str(e)}"
            logger.error(f"❌ {error_msg}", exc_info=True)
            return False, error_msg

    def export_dashboard_to_pdf(
        self,
        metrics: Dict,
        chart_images: Optional[List[bytes]] = None,
        filepath: str = None,
        date_range: str = None
    ) -> Tuple[bool, str]:
        """
        Exportar dashboard a PDF profesional con estilo corporativo
        """
        if not HAS_REPORTLAB:
            return False, "reportlab no instalado. Instala con: pip install reportlab"

        try:
            doc = SimpleDocTemplate(
                filepath,
                pagesize=letter,
                rightMargin=0.6*inch,
                leftMargin=0.6*inch,
                topMargin=0.6*inch,
                bottomMargin=0.5*inch
            )
            
            story = []
            
            # Header
            story.append(Paragraph(
                "<b>REPORTE DE VENTAS</b>", 
                ParagraphStyle('Title', fontSize=18, textColor=self.COLOR_PRIMARY, fontName='Helvetica-Bold')
            ))
            story.append(Spacer(1, 0.15*inch))
            story.append(Paragraph(
                f"Empresa: Disfruleg | {datetime.now().strftime('%d/%m/%Y %H:%M')}", 
                ParagraphStyle('Meta', fontSize=9, textColor=self.COLOR_TEXT)
            ))
            if date_range:
                story.append(Paragraph(
                    f"Período: {date_range}", 
                    ParagraphStyle('Meta', fontSize=8, textColor=self.COLOR_TEXT)
                ))
            story.append(Spacer(1, 0.2*inch))
            
            # KPIs
            kpi_data = [
                [
                    Paragraph(f"${metrics.get('total_ingresos', 0):,.2f}", 
                        ParagraphStyle('KPI', fontSize=14, textColor=self.COLOR_PRIMARY, fontName='Helvetica-Bold', alignment=TA_CENTER)),
                    Paragraph(f"${metrics.get('total_costos', 0):,.2f}", 
                        ParagraphStyle('KPI', fontSize=14, textColor=self.COLOR_PRIMARY, fontName='Helvetica-Bold', alignment=TA_CENTER)),
                    Paragraph(f"${metrics.get('total_ganancia', 0):,.2f}", 
                        ParagraphStyle('KPI', fontSize=14, textColor=self.COLOR_PRIMARY, fontName='Helvetica-Bold', alignment=TA_CENTER)),
                    Paragraph(f"{metrics.get('margen_promedio', 0):.1f}%", 
                        ParagraphStyle('KPI', fontSize=14, textColor=self.COLOR_PRIMARY, fontName='Helvetica-Bold', alignment=TA_CENTER)),
                ],
                [
                    Paragraph("INGRESOS", ParagraphStyle('Label', fontSize=8, textColor=self.COLOR_TEXT, alignment=TA_CENTER)),
                    Paragraph("COSTOS", ParagraphStyle('Label', fontSize=8, textColor=self.COLOR_TEXT, alignment=TA_CENTER)),
                    Paragraph("GANANCIA", ParagraphStyle('Label', fontSize=8, textColor=self.COLOR_TEXT, alignment=TA_CENTER)),
                    Paragraph("MARGEN %", ParagraphStyle('Label', fontSize=8, textColor=self.COLOR_TEXT, alignment=TA_CENTER)),
                ]
            ]
            
            kpi_table = Table(kpi_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
            kpi_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.COLOR_ACCENT),
                ('BACKGROUND', (0, 1), (-1, 1), self.COLOR_BG),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 1), (-1, 1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, 1), 6),
                ('GRID', (0, 0), (-1, -1), 1, self.COLOR_BG),
                ('LINEBELOW', (0, 0), (-1, 0), 2, self.COLOR_PRIMARY),
            ]))
            story.append(kpi_table)
            story.append(Spacer(1, 0.3*inch))
            
            # Footer
            story.append(Paragraph(
                "Reporte generado automáticamente por C.H.U.M.I, desarrollado por Ubicuo Studio",
                ParagraphStyle('Footer', fontSize=7, textColor=colors.grey, fontName='Helvetica-Oblique', alignment=TA_CENTER)
            ))
            
            doc.build(story)
            logger.info(f"✅ Dashboard PDF exported: {filepath}")
            return True, filepath
            
        except Exception as e:
            error_msg = f"Error al exportar PDF: {str(e)}"
            logger.error(f"❌ {error_msg}", exc_info=True)
            return False, error_msg

    def export_products_report_pdf(
        self,
        products: List[Dict],
        filepath: str,
        date_range: str = None
    ) -> Tuple[bool, str]:
        """
        Exportar reporte de productos a PDF profesional
        """
        if not HAS_REPORTLAB:
            return False, "reportlab no instalado"

        try:
            doc = SimpleDocTemplate(
                filepath,
                pagesize=letter,
                rightMargin=0.6*inch,
                leftMargin=0.6*inch,
                topMargin=0.6*inch,
                bottomMargin=0.5*inch
            )
            
            story = []
            
            # Header
            story.append(Paragraph(
                "<b>REPORTE - PRODUCTOS</b>", 
                ParagraphStyle('Title', fontSize=18, textColor=self.COLOR_PRIMARY, fontName='Helvetica-Bold')
            ))
            story.append(Spacer(1, 0.15*inch))
            story.append(Paragraph(
                f"Empresa: Disfruleg | {datetime.now().strftime('%d/%m/%Y %H:%M')}", 
                ParagraphStyle('Meta', fontSize=9, textColor=self.COLOR_TEXT)
            ))
            if date_range:
                story.append(Paragraph(
                    f"Período: {date_range}", 
                    ParagraphStyle('Meta', fontSize=8, textColor=self.COLOR_TEXT)
                ))
            story.append(Spacer(1, 0.2*inch))
            
            # Título de tabla centrado
            story.append(Paragraph(
                "Top 10 Productos con Mayor Rentabilidad",
                ParagraphStyle('SectionTitle', fontSize=12, textColor=self.COLOR_PRIMARY, fontName='Helvetica-Bold', alignment=TA_CENTER)
            ))
            story.append(Spacer(1, 0.15*inch))
            
            # Tabla de productos
            data = [["Producto", "Cantidad", "Ganancia", "% Contrib."]]
            total_ganancia = sum([p.get('ganancia_total', 0) for p in products])
            
            for p in products[:20]:
                porcentaje = (p.get('ganancia_total', 0) / total_ganancia * 100) if total_ganancia > 0 else 0
                data.append([
                    Paragraph(str(p.get('nombre_producto', 'N/A'))[:30], 
                        ParagraphStyle('Cell', fontSize=8, textColor=self.COLOR_TEXT, alignment=TA_LEFT)),
                    Paragraph(f"{int(p.get('cantidad_vendida', 0))}", 
                        ParagraphStyle('Cell', fontSize=8, textColor=self.COLOR_TEXT, alignment=TA_RIGHT)),
                    Paragraph(f"${p.get('ganancia_total', 0):,.2f}", 
                        ParagraphStyle('Cell', fontSize=8, textColor=self.COLOR_TEXT, alignment=TA_RIGHT)),
                    Paragraph(f"{porcentaje:.1f}%", 
                        ParagraphStyle('Cell', fontSize=8, textColor=self.COLOR_TEXT, alignment=TA_RIGHT)),
                ])
            
            table = Table(data, colWidths=[3.2*inch, 1*inch, 1*inch, 0.8*inch])
            
            style_list = [
                ('BACKGROUND', (0, 0), (-1, 0), self.COLOR_PRIMARY),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, self.COLOR_BG),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
            
            # Alternar colores de fondo
            for i in range(1, len(data)):
                if (i - 1) % 2 == 0:
                    style_list.append(('BACKGROUND', (0, i), (-1, i), self.COLOR_ACCENT))
                else:
                    style_list.append(('BACKGROUND', (0, i), (-1, i), colors.white))
            
            table.setStyle(TableStyle(style_list))
            story.append(table)
            story.append(Spacer(1, 0.3*inch))
            
            # Footer
            story.append(Paragraph(
                "Reporte generado automáticamente por C.H.U.M.I, desarrollado por Ubicuo Studio",
                ParagraphStyle('Footer', fontSize=7, textColor=colors.grey, fontName='Helvetica-Oblique', alignment=TA_CENTER)
            ))
            
            doc.build(story)
            logger.info(f"✅ Products PDF exported: {filepath}")
            return True, filepath
            
        except Exception as e:
            error_msg = f"Error PDF productos: {str(e)}"
            logger.error(error_msg)
            return False, error_msg

    def export_clients_report_pdf(
        self,
        clients: List[Dict],
        filepath: str,
        date_range: str = None
    ) -> Tuple[bool, str]:
        """
        Exportar reporte de clientes a PDF profesional
        """
        if not HAS_REPORTLAB:
            return False, "reportlab no instalado"

        try:
            doc = SimpleDocTemplate(
                filepath,
                pagesize=letter,
                rightMargin=0.6*inch,
                leftMargin=0.6*inch,
                topMargin=0.6*inch,
                bottomMargin=0.5*inch
            )
            
            story = []
            
            # Header
            story.append(Paragraph(
                "<b>REPORTE - CLIENTES</b>", 
                ParagraphStyle('Title', fontSize=18, textColor=self.COLOR_PRIMARY, fontName='Helvetica-Bold')
            ))
            story.append(Spacer(1, 0.15*inch))
            story.append(Paragraph(
                f"Empresa: Disfruleg | {datetime.now().strftime('%d/%m/%Y %H:%M')}", 
                ParagraphStyle('Meta', fontSize=9, textColor=self.COLOR_TEXT)
            ))
            if date_range:
                story.append(Paragraph(
                    f"Período: {date_range}", 
                    ParagraphStyle('Meta', fontSize=8, textColor=self.COLOR_TEXT)
                ))
            story.append(Spacer(1, 0.2*inch))
            
            # Título de tabla centrado
            story.append(Paragraph(
                "Ranking de Clientes por Gasto Total",
                ParagraphStyle('SectionTitle', fontSize=12, textColor=self.COLOR_PRIMARY, fontName='Helvetica-Bold', alignment=TA_CENTER)
            ))
            story.append(Spacer(1, 0.15*inch))
            
            # Tabla de clientes
            data = [["Cliente", "Compras", "Gasto Total", "Ticket Promedio"]]
            
            for c in clients[:20]:
                data.append([
                    Paragraph(str(c.get('nombre_cliente', 'N/A'))[:25], 
                        ParagraphStyle('Cell', fontSize=8, textColor=self.COLOR_TEXT, alignment=TA_LEFT)),
                    Paragraph(f"{int(c.get('numero_compras', 0))}", 
                        ParagraphStyle('Cell', fontSize=8, textColor=self.COLOR_TEXT, alignment=TA_RIGHT)),
                    Paragraph(f"${c.get('gasto_total', 0):,.2f}", 
                        ParagraphStyle('Cell', fontSize=8, textColor=self.COLOR_TEXT, alignment=TA_RIGHT)),
                    Paragraph(f"${c.get('ticket_promedio', 0):,.2f}", 
                        ParagraphStyle('Cell', fontSize=8, textColor=self.COLOR_TEXT, alignment=TA_RIGHT)),
                ])
            
            table = Table(data, colWidths=[2.4*inch, 1*inch, 1.4*inch, 1.4*inch])
            
            style_list = [
                ('BACKGROUND', (0, 0), (-1, 0), self.COLOR_PRIMARY),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, self.COLOR_BG),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
            
            # Alternar colores de fondo
            for i in range(1, len(data)):
                if (i - 1) % 2 == 0:
                    style_list.append(('BACKGROUND', (0, i), (-1, i), self.COLOR_ACCENT))
                else:
                    style_list.append(('BACKGROUND', (0, i), (-1, i), colors.white))
            
            table.setStyle(TableStyle(style_list))
            story.append(table)
            story.append(Spacer(1, 0.3*inch))
            
            # Footer
            story.append(Paragraph(
                "Reporte generado automáticamente por C.H.U.M.I, desarrollado por Ubicuo Studio",
                ParagraphStyle('Footer', fontSize=7, textColor=colors.grey, fontName='Helvetica-Oblique', alignment=TA_CENTER)
            ))
            
            doc.build(story)
            logger.info(f"✅ Clients PDF exported: {filepath}")
            return True, filepath
            
        except Exception as e:
            error_msg = f"Error PDF clientes: {str(e)}"
            logger.error(error_msg)
            return False, error_msg

    def _format_excel_workbook(self, filepath: str):
        """
        Aplicar formato profesional a workbook de Excel
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = load_workbook(filepath)

            # Estilo de encabezado - Verde corporativo
            header_fill = PatternFill(start_color="2D5A27", end_color="2D5A27", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)

            for sheet in wb.worksheets:
                # Formato de fila de encabezado
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Auto-ajustar ancho de columnas
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width

            wb.save(filepath)
            logger.debug("Excel formatting applied")

        except Exception as e:
            logger.warning(f"Could not apply Excel formatting: {e}")

    def _format_date(self, date_value) -> str:
        """
        Formatear valor de fecha para exportación
        """
        if date_value is None:
            return "N/A"

        if isinstance(date_value, str):
            return date_value

        if hasattr(date_value, 'strftime'):
            return date_value.strftime('%Y-%m-%d')

        return str(date_value)