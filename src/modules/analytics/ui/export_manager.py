# -*- coding: utf-8 -*-
"""
DISFRULEG - Export Manager (VERSIÓN DEFINITIVA)
Exportación a PDF y Excel - SIN ERRORES
"""
import os
import json
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from tkinter import filedialog
import logging

logger = logging.getLogger(__name__)

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExportManager:
    """Gestor de exportación a PDF y Excel con configuración persistente"""

    CONFIG_FILE = os.path.join(os.path.expanduser("~"), ".disfruleg_export_config.json")

    def __init__(self, data_manager, output_dir: str = None):
        self.data_manager = data_manager
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Load saved directory or use provided/default
        if output_dir:
            self.output_dir = output_dir
        else:
            self.output_dir = self._load_saved_directory()

        # Ensure directory exists
        try:
            os.makedirs(self.output_dir, exist_ok=True)
        except:
            self.output_dir = os.path.expanduser("~/Downloads")
            os.makedirs(self.output_dir, exist_ok=True)

    def _load_saved_directory(self) -> str:
        """Load previously saved export directory from config file"""
        try:
            if os.path.exists(self.CONFIG_FILE):
                with open(self.CONFIG_FILE, 'r') as f:
                    config = json.load(f)
                    saved_dir = config.get('export_directory')
                    if saved_dir and os.path.isdir(saved_dir):
                        logger.info(f"Loaded saved export directory: {saved_dir}")
                        return saved_dir
        except Exception as e:
            logger.warning(f"Could not load saved directory: {e}")

        # If no saved directory or error, return default
        return os.path.expanduser("~/Downloads")

    def _save_directory(self, directory: str):
        """Save export directory to config file for future use"""
        try:
            config = {'export_directory': directory}
            with open(self.CONFIG_FILE, 'w') as f:
                json.dump(config, f)
            logger.info(f"Saved export directory: {directory}")
        except Exception as e:
            logger.warning(f"Could not save directory config: {e}")

    def ask_and_set_output_directory(self, parent=None) -> bool:
        """
        Ask user to select export directory and save it as default

        Returns:
            bool: True if directory was selected, False if cancelled
        """
        selected_dir = filedialog.askdirectory(
            parent=parent,
            title="Seleccionar carpeta para exportaciones",
            initialdir=self.output_dir,
            mustexist=True
        )

        if selected_dir:
            self.output_dir = selected_dir
            self._save_directory(selected_dir)

            # Ensure directory exists
            try:
                os.makedirs(self.output_dir, exist_ok=True)
            except:
                pass

            return True

        return False

    def get_output_directory(self) -> str:
        """Get current output directory"""
        return self.output_dir
    
    def export_dashboard_pdf(self, dashboard_data: Dict) -> Tuple[bool, str]:
        """Exportar dashboard a PDF con formato profesional"""
        if not HAS_REPORTLAB:
            return False, "reportlab no instalado"

        try:
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT
            
            # Colores corporativos
            COLOR_PRIMARY = colors.HexColor("#2D5A27")
            COLOR_TEXT = colors.HexColor("#333333")
            COLOR_BG = colors.HexColor("#F4F4F4")
            COLOR_ACCENT = colors.HexColor("#E8F0E5")
            
            filename = f"Dashboard_{self.timestamp}.pdf"
            filepath = os.path.join(self.output_dir, filename)

            doc = SimpleDocTemplate(
                filepath,
                pagesize=letter,
                rightMargin=0.6*inch,
                leftMargin=0.6*inch,
                topMargin=0.6*inch,
                bottomMargin=0.5*inch
            )
            
            story = []
            
            # Header
            story.append(Paragraph(
                "<b>REPORTE DE VENTAS</b>",
                ParagraphStyle('Title', fontSize=20, textColor=COLOR_PRIMARY, fontName='Helvetica-Bold')
            ))
            story.append(Spacer(1, 0.1*inch))
            story.append(Paragraph(
                "Empresa: Disfruleg",
                ParagraphStyle('Subtitle', fontSize=11, textColor=COLOR_TEXT, fontName='Helvetica')
            ))
            story.append(Spacer(1, 0.15*inch))
            
            # Meta información alineada a derecha
            meta_style = ParagraphStyle('Meta', fontSize=9, textColor=COLOR_TEXT, alignment=TA_RIGHT)
            date_range = dashboard_data.get('date_range', 'N/A')
            meta_table = Table([
                [
                    "",
                    Paragraph(f"{datetime.now().strftime('%d/%m/%Y %H:%M')}<br/>Período: {date_range}", meta_style)
                ]
            ], colWidths=[4*inch, 2.8*inch])
            meta_table.setStyle(TableStyle([
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                ('TOPPADDING', (0, 0), (-1, 0), 0),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 0),
            ]))
            story.append(meta_table)
            story.append(Spacer(1, 0.2*inch))
            
            # Línea divisora
            divider = Table([['']],  colWidths=[7.2*inch])
            divider.setStyle(TableStyle([
                ('LINEBELOW', (0, 0), (0, 0), 1.5, COLOR_PRIMARY),
            ]))
            story.append(divider)
            story.append(Spacer(1, 0.2*inch))
            
            # KPIs
            if 'kpis' in dashboard_data and dashboard_data['kpis']:
                kpis = dashboard_data['kpis']
                kpi_data = [
                    [
                        Paragraph(f"${kpis.get('total_ingresos', 0):,.2f}", 
                            ParagraphStyle('KPI', fontSize=14, textColor=COLOR_PRIMARY, fontName='Helvetica-Bold', alignment=TA_CENTER)),
                        Paragraph(f"${kpis.get('total_costos', 0):,.2f}", 
                            ParagraphStyle('KPI', fontSize=14, textColor=COLOR_PRIMARY, fontName='Helvetica-Bold', alignment=TA_CENTER)),
                        Paragraph(f"${kpis.get('total_ganancia', 0):,.2f}", 
                            ParagraphStyle('KPI', fontSize=14, textColor=COLOR_PRIMARY, fontName='Helvetica-Bold', alignment=TA_CENTER)),
                        Paragraph(f"{kpis.get('margen_promedio', 0):.1f}%", 
                            ParagraphStyle('KPI', fontSize=14, textColor=COLOR_PRIMARY, fontName='Helvetica-Bold', alignment=TA_CENTER)),
                    ],
                    [
                        Paragraph("INGRESOS", ParagraphStyle('Label', fontSize=8, textColor=COLOR_TEXT, alignment=TA_CENTER)),
                        Paragraph("COSTOS", ParagraphStyle('Label', fontSize=8, textColor=COLOR_TEXT, alignment=TA_CENTER)),
                        Paragraph("GANANCIA NETA", ParagraphStyle('Label', fontSize=8, textColor=COLOR_TEXT, alignment=TA_CENTER)),
                        Paragraph("MARGEN %", ParagraphStyle('Label', fontSize=8, textColor=COLOR_TEXT, alignment=TA_CENTER)),
                    ]
                ]
                
                kpi_table = Table(kpi_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
                kpi_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), COLOR_ACCENT),
                    ('BACKGROUND', (0, 1), (-1, 1), COLOR_BG),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                    ('TOPPADDING', (0, 1), (-1, 1), 6),
                    ('BOTTOMPADDING', (0, 1), (-1, 1), 6),
                    ('GRID', (0, 0), (-1, -1), 1, COLOR_BG),
                    ('LINEBELOW', (0, 0), (-1, 0), 2, COLOR_PRIMARY),
                ]))
                story.append(kpi_table)
                story.append(Spacer(1, 0.3*inch))
            
            # Título centrado
            story.append(Paragraph(
                "Top 10 Productos con Mayor Rentabilidad",
                ParagraphStyle('SectionTitle', fontSize=12, textColor=COLOR_PRIMARY, fontName='Helvetica-Bold', alignment=TA_CENTER)
            ))
            story.append(Spacer(1, 0.15*inch))
            
            # Tabla de productos
            if 'top_products' in dashboard_data and dashboard_data['top_products']:
                data = [["Producto", "Cantidad", "Ganancia", "% Contrib."]]
                products = dashboard_data.get('top_products', [])
                total_ganancia = sum([p.get('ganancia_total', 0) for p in products])
                
                for p in products[:10]:
                    porcentaje = (p.get('ganancia_total', 0) / total_ganancia * 100) if total_ganancia > 0 else 0
                    data.append([
                        Paragraph(str(p.get('nombre_producto', 'N/A'))[:28], 
                            ParagraphStyle('Cell', fontSize=8, textColor=COLOR_TEXT, alignment=TA_CENTER)),
                        Paragraph(f"{int(p.get('cantidad_vendida', 0))}", 
                            ParagraphStyle('Cell', fontSize=8, textColor=COLOR_TEXT, alignment=TA_CENTER)),
                        Paragraph(f"${p.get('ganancia_total', 0):,.2f}", 
                            ParagraphStyle('Cell', fontSize=8, textColor=COLOR_TEXT, alignment=TA_CENTER)),
                        Paragraph(f"{porcentaje:.1f}%", 
                            ParagraphStyle('Cell', fontSize=8, textColor=COLOR_TEXT, alignment=TA_CENTER)),
                    ])
                
                table = Table(data, colWidths=[3.2*inch, 1*inch, 1*inch, 0.8*inch])
                
                style_list = [
                    ('BACKGROUND', (0, 0), (-1, 0), COLOR_PRIMARY),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('TOPPADDING', (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, COLOR_BG),
                    ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ]
                
                for i in range(1, len(data)):
                    if (i - 1) % 2 == 0:
                        style_list.append(('BACKGROUND', (0, i), (-1, i), COLOR_ACCENT))
                    else:
                        style_list.append(('BACKGROUND', (0, i), (-1, i), colors.white))
                
                table.setStyle(TableStyle(style_list))
                story.append(table)
                story.append(Spacer(1, 0.3*inch))
            
            # Footer
            story.append(Paragraph(
                "Reporte generado automáticamente por C.H.U.M.I, desarrollado por Ubicuo Studio",
                ParagraphStyle('Footer', fontSize=7, textColor=colors.grey, fontName='Helvetica-Oblique', alignment=TA_CENTER)
            ))
            
            doc.build(story)
            logger.info(f"✅ Dashboard PDF exported: {filepath}")
            return True, filepath
        
        except Exception as e:
            logger.error(f"❌ Error PDF: {e}")
            return False, str(e)
    
    def export_products_report_pdf(self, products: List[Dict]) -> Tuple[bool, str]:
        """Exportar reporte de productos a PDF profesional"""
        if not HAS_REPORTLAB:
            return False, "reportlab no instalado"
        
        try:
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT
            
            COLOR_PRIMARY = colors.HexColor("#2D5A27")
            COLOR_TEXT = colors.HexColor("#333333")
            COLOR_BG = colors.HexColor("#F4F4F4")
            COLOR_ACCENT = colors.HexColor("#E8F0E5")
            
            filename = f"Productos_{self.timestamp}.pdf"
            filepath = os.path.join(self.output_dir, filename)
            
            doc = SimpleDocTemplate(
                filepath,
                pagesize=letter,
                rightMargin=0.6*inch,
                leftMargin=0.6*inch,
                topMargin=0.6*inch,
                bottomMargin=0.5*inch
            )
            
            story = []
            
            # Header
            story.append(Paragraph(
                "<b>REPORTE - PRODUCTOS</b>",
                ParagraphStyle('Title', fontSize=20, textColor=COLOR_PRIMARY, fontName='Helvetica-Bold')
            ))
            story.append(Spacer(1, 0.1*inch))
            story.append(Paragraph(
                "Empresa: Disfruleg",
                ParagraphStyle('Subtitle', fontSize=11, textColor=COLOR_TEXT, fontName='Helvetica')
            ))
            story.append(Spacer(1, 0.2*inch))
            
            # Meta
            meta_style = ParagraphStyle('Meta', fontSize=9, textColor=COLOR_TEXT, alignment=TA_RIGHT)
            meta_table = Table([
                [
                    "",
                    Paragraph(f"{datetime.now().strftime('%d/%m/%Y %H:%M')}", meta_style)
                ]
            ], colWidths=[4*inch, 2.8*inch])
            meta_table.setStyle(TableStyle([
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                ('TOPPADDING', (0, 0), (-1, 0), 0),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 0),
            ]))
            story.append(meta_table)
            story.append(Spacer(1, 0.1*inch))
            
            # Tabla
            data = [["Producto", "Cantidad", "Ganancia", "%"]]
            total_ganancia = sum([p.get('ganancia_total', 0) for p in products])
            
            for p in products[:20]:
                porcentaje = (p.get('ganancia_total', 0) / total_ganancia * 100) if total_ganancia > 0 else 0
                data.append([
                    Paragraph(str(p.get('nombre_producto', 'N/A'))[:28], 
                        ParagraphStyle('Cell', fontSize=8, textColor=COLOR_TEXT)),
                    Paragraph(f"{int(p.get('cantidad_vendida', 0))}", 
                        ParagraphStyle('Cell', fontSize=8, textColor=COLOR_TEXT, alignment=TA_RIGHT)),
                    Paragraph(f"${p.get('ganancia_total', 0):,.2f}", 
                        ParagraphStyle('Cell', fontSize=8, textColor=COLOR_TEXT, alignment=TA_RIGHT)),
                    Paragraph(f"{porcentaje:.1f}%", 
                        ParagraphStyle('Cell', fontSize=8, textColor=COLOR_TEXT, alignment=TA_RIGHT)),
                ])
            
            table = Table(data, colWidths=[3.2*inch, 1*inch, 1*inch, 0.8*inch])
            
            style_list = [
                ('BACKGROUND', (0, 0), (-1, 0), COLOR_PRIMARY),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, COLOR_BG),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ]
            
            for i in range(1, len(data)):
                if (i - 1) % 2 == 0:
                    style_list.append(('BACKGROUND', (0, i), (-1, i), COLOR_ACCENT))
                else:
                    style_list.append(('BACKGROUND', (0, i), (-1, i), colors.white))
            
            table.setStyle(TableStyle(style_list))
            story.append(table)
            story.append(Spacer(1, 0.3*inch))
            
            # Footer
            story.append(Paragraph(
                "Reporte generado automáticamente por C.H.U.M.I, desarrollado por Ubicuo Studio",
                ParagraphStyle('Footer', fontSize=7, textColor=colors.grey, fontName='Helvetica-Oblique', alignment=TA_CENTER)
            ))
            
            doc.build(story)
            logger.info(f"✅ Products PDF exported: {filepath}")
            return True, filepath
        
        except Exception as e:
            logger.error(f"❌ Error PDF productos: {e}")
            return False, str(e)
    
    def export_clients_report_pdf(self, clients: List[Dict]) -> Tuple[bool, str]:
        """Exportar reporte de clientes a PDF profesional"""
        if not HAS_REPORTLAB:
            return False, "reportlab no instalado"
        
        try:
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT
            
            COLOR_PRIMARY = colors.HexColor("#2D5A27")
            COLOR_TEXT = colors.HexColor("#333333")
            COLOR_BG = colors.HexColor("#F4F4F4")
            COLOR_ACCENT = colors.HexColor("#E8F0E5")
            
            filename = f"Clientes_{self.timestamp}.pdf"
            filepath = os.path.join(self.output_dir, filename)
            
            doc = SimpleDocTemplate(
                filepath,
                pagesize=letter,
                rightMargin=0.6*inch,
                leftMargin=0.6*inch,
                topMargin=0.6*inch,
                bottomMargin=0.5*inch
            )
            
            story = []
            
            # Header
            story.append(Paragraph(
                "<b>REPORTE - CLIENTES</b>",
                ParagraphStyle('Title', fontSize=20, textColor=COLOR_PRIMARY, fontName='Helvetica-Bold')
            ))
            story.append(Spacer(1, 0.1*inch))
            story.append(Paragraph(
                "Empresa: Disfruleg",
                ParagraphStyle('Subtitle', fontSize=11, textColor=COLOR_TEXT, fontName='Helvetica')
            ))
            story.append(Spacer(1, 0.2*inch))
            
            # Meta
            meta_style = ParagraphStyle('Meta', fontSize=9, textColor=COLOR_TEXT, alignment=TA_RIGHT)
            meta_table = Table([
                [
                    "",
                    Paragraph(f"{datetime.now().strftime('%d/%m/%Y %H:%M')}", meta_style)
                ]
            ], colWidths=[4*inch, 2.8*inch])
            meta_table.setStyle(TableStyle([
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                ('TOPPADDING', (0, 0), (-1, 0), 0),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 0),
            ]))
            story.append(meta_table)
            story.append(Spacer(1, 0.1*inch))
            
            # Tabla
            data = [["Cliente", "Compras", "Gasto", "Ticket"]]
            
            for c in clients[:20]:
                data.append([
                    Paragraph(str(c.get('nombre_cliente', 'N/A'))[:25], 
                        ParagraphStyle('Cell', fontSize=8, textColor=COLOR_TEXT)),
                    Paragraph(str(c.get('numero_compras', 0)), 
                        ParagraphStyle('Cell', fontSize=8, textColor=COLOR_TEXT, alignment=TA_RIGHT)),
                    Paragraph(f"${c.get('gasto_total', 0):,.2f}", 
                        ParagraphStyle('Cell', fontSize=8, textColor=COLOR_TEXT, alignment=TA_RIGHT)),
                    Paragraph(f"${c.get('ticket_promedio', 0):,.2f}", 
                        ParagraphStyle('Cell', fontSize=8, textColor=COLOR_TEXT, alignment=TA_RIGHT)),
                ])
            
            table = Table(data, colWidths=[2.5*inch, 1*inch, 1.4*inch, 1.4*inch])
            
            style_list = [
                ('BACKGROUND', (0, 0), (-1, 0), COLOR_PRIMARY),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, COLOR_BG),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ]
            
            for i in range(1, len(data)):
                if (i - 1) % 2 == 0:
                    style_list.append(('BACKGROUND', (0, i), (-1, i), COLOR_ACCENT))
                else:
                    style_list.append(('BACKGROUND', (0, i), (-1, i), colors.white))
            
            table.setStyle(TableStyle(style_list))
            story.append(table)
            story.append(Spacer(1, 0.3*inch))
            
            # Footer
            story.append(Paragraph(
                "Reporte generado automáticamente por C.H.U.M.I, desarrollado por Ubicuo Studio",
                ParagraphStyle('Footer', fontSize=7, textColor=colors.grey, fontName='Helvetica-Oblique', alignment=TA_CENTER)
            ))
            
            doc.build(story)
            logger.info(f"✅ Clients PDF exported: {filepath}")
            return True, filepath
        
        except Exception as e:
            logger.error(f"❌ Error PDF clientes: {e}")
            return False, str(e)
    
    def export_dashboard_excel(self, dashboard_data: Dict) -> Tuple[bool, str]:
        """Exportar dashboard a Excel"""
        if not HAS_OPENPYXL:
            return False, "openpyxl no instalado"

        try:
            filename = f"Dashboard_{self.timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)

            wb = Workbook()
            wb.remove(wb.active)

            # KPIs
            if 'kpis' in dashboard_data and dashboard_data['kpis']:
                ws = wb.create_sheet("KPIs")
                kpis = dashboard_data['kpis']

                # Add date range header if available
                row = 1
                if 'date_range' in dashboard_data:
                    ws['A1'] = "Periodo"
                    ws['B1'] = dashboard_data['date_range']
                    row = 3

                ws[f'A{row}'] = "Métrica"
                ws[f'B{row}'] = "Valor"
                ws[f'A{row+1}'] = "Ganancia Total"
                ws[f'B{row+1}'] = kpis.get('total_ganancia', 0)
                ws[f'A{row+2}'] = "Ingresos"
                ws[f'B{row+2}'] = kpis.get('total_ingresos', 0)
                ws[f'A{row+3}'] = "Costos"
                ws[f'B{row+3}'] = kpis.get('total_costos', 0)
                self._format_excel(ws)
            
            # Productos
            if 'top_products' in dashboard_data and dashboard_data['top_products']:
                ws = wb.create_sheet("Productos")
                ws['A1'] = "Producto"
                ws['B1'] = "Cantidad"
                ws['C1'] = "Ganancia"
                ws['D1'] = "Margen %"
                row = 2
                for p in dashboard_data['top_products'][:20]:
                    ws[f'A{row}'] = p.get('nombre_producto', 'N/A')
                    ws[f'B{row}'] = p.get('cantidad_vendida', 0)
                    ws[f'C{row}'] = p.get('ganancia_total', 0)
                    ws[f'D{row}'] = p.get('margen_ganancia_porcentaje', 0)
                    row += 1
                self._format_excel(ws)
            
            # Clientes
            if 'top_clients' in dashboard_data and dashboard_data['top_clients']:
                ws = wb.create_sheet("Clientes")
                ws['A1'] = "Cliente"
                ws['B1'] = "Compras"
                ws['C1'] = "Gasto Total"
                ws['D1'] = "Ticket"
                row = 2
                for c in dashboard_data['top_clients'][:20]:
                    ws[f'A{row}'] = c.get('nombre_cliente', 'N/A')
                    ws[f'B{row}'] = c.get('numero_compras', 0)
                    ws[f'C{row}'] = c.get('gasto_total', 0)
                    ws[f'D{row}'] = c.get('ticket_promedio', 0)
                    row += 1
                self._format_excel(ws)
            
            # Grupos
            if 'group_summary' in dashboard_data and dashboard_data['group_summary']:
                ws = wb.create_sheet("Grupos")
                ws['A1'] = "Grupo"
                ws['B1'] = "Ingresos"
                ws['C1'] = "Ventas"
                row = 2
                for g in dashboard_data['group_summary']:
                    ws[f'A{row}'] = g.get('nombre_grupo', 'N/A')
                    ws[f'B{row}'] = g.get('ingresos', 0)
                    ws[f'C{row}'] = g.get('ganancias', 0)
                    row += 1
                self._format_excel(ws)
            
            # Crear hoja mínima si está vacío
            if len(wb.sheetnames) == 0:
                ws = wb.create_sheet("Datos")
                ws['A1'] = "No hay datos"
            
            wb.save(filepath)
            logger.info(f"Excel: {filepath}")
            return True, filepath
        
        except Exception as e:
            logger.error(f"Error Excel: {e}")
            return False, str(e)
    
    def export_section_excel(self, section_name: str, data: List[Dict]) -> Tuple[bool, str]:
        """Exportar sección a Excel"""
        if not HAS_OPENPYXL:
            return False, "openpyxl no instalado"
        
        try:
            filename = f"{section_name}_{self.timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            
            wb = Workbook()
            ws = wb.active
            ws.title = section_name
            
            if data:
                headers = list(data[0].keys())
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                for row, item in enumerate(data, 2):
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=row, column=col, value=item.get(header, ''))
            
            self._format_excel(ws)
            wb.save(filepath)
            return True, filepath
        
        except Exception as e:
            logger.error(f"Error Excel sección: {e}")
            return False, str(e)
    
    def _format_excel(self, ws):
        """Formatear hoja Excel"""
        header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.border = border
        
        for col_num, col in enumerate(ws.iter_cols(), 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15


class ChartExporter:
    """Exportador de gráficos"""
    
    @staticmethod
    def save_chart_image(fig, output_path: str) -> bool:
        """Guardar figura matplotlib"""
        try:
            fig.savefig(output_path, dpi=150, bbox_inches='tight')
            return True
        except:
            return False
    
    @staticmethod
    def create_sales_chart(dates: List, values: List) -> str:
        """Crear gráfico de ventas"""
        try:
            import matplotlib.pyplot as plt
            fig, ax = plt.subplots(figsize=(12, 6))
            ax.plot(dates, values, marker='o', linewidth=2)
            ax.fill_between(range(len(dates)), values, alpha=0.3)
            ax.set_xlabel('Fecha')
            ax.set_ylabel('Ventas')
            ax.set_title('Últimos 30 días')
            ax.grid(True, alpha=0.3)
            
            output_path = f"/tmp/chart_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            ChartExporter.save_chart_image(fig, output_path)
            plt.close(fig)
            return output_path
        except:
            return None